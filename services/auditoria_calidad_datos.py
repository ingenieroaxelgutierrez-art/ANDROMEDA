# ============================================================
# AUDITORÍA DE CALIDAD DE DATOS - Triple Validación Odoo
# ============================================================
# Módulo de detección de incoherencias en toda la base de datos:
# 1. Estado vs Vínculo (Procesos Huérfanos)
# 2. Tiempo de Vida / SLA (Registros Zombi)
# 3. Cálculo de Incertidumbre (% datos basura vs confiables)
# Genera Excel con hallazgos. Solo lectura.
# ============================================================

import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field
from collections import defaultdict

try:
    import pandas as pd
    PANDAS_DISPONIBLE = True
except ImportError:
    PANDAS_DISPONIBLE = False

from app.logging_config import get_logger
logger = get_logger("services.auditoria_calidad_datos")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False


# ============================================================
# ESTRUCTURAS DE DATOS
# ============================================================

@dataclass
class HallazgoCalidad:
    """Un hallazgo individual de la auditoría de calidad."""
    categoria: str          # incompleto, inconsistente
    severidad: str          # critico, alto, medio, bajo
    modelo_odoo: str        # sale.order, account.move, etc.
    registro_id: int
    registro_nombre: str
    descripcion: str
    detalle: str
    accion_sugerida: str
    empresa: str = ''
    unidad_operativa: str = ''
    usuario_creador: str = ''
    datos_extra: Dict = field(default_factory=dict)
    fecha_deteccion: datetime = field(default_factory=datetime.now)

    def to_dict(self) -> Dict:
        return {
            'Categoría': self.categoria,
            'Severidad': self.severidad,
            'Modelo': self.modelo_odoo,
            'ID': self.registro_id,
            'Referencia': self.registro_nombre,
            'Descripción': self.descripcion,
            'Detalle': self.detalle,
            'Acción Sugerida': self.accion_sugerida,
            'Empresa': self.empresa,
            'Unidad Operativa': self.unidad_operativa,
            'Usuario Creador': self.usuario_creador,
            'Fecha Detección': self.fecha_deteccion.strftime('%Y-%m-%d %H:%M'),
            **{k: str(v) for k, v in self.datos_extra.items()}
        }


@dataclass
class ResultadoCalidadDatos:
    """Resultado completo de la auditoría de calidad de datos."""
    fecha_ejecucion: datetime
    total_registros_analizados: int
    hallazgos: List[HallazgoCalidad]
    resumen_por_categoria: Dict[str, int]
    resumen_por_severidad: Dict[str, int]
    resumen_por_modelo: Dict[str, int]
    porcentaje_datos_confiables: float
    porcentaje_datos_basura: float
    indice_incertidumbre: float       # 0-100
    resumen_por_empresa: Dict[str, int] = field(default_factory=dict)
    resumen_por_unidad_operativa: Dict[str, int] = field(default_factory=dict)
    top_usuarios_problemas: List = field(default_factory=list)
    ruta_excel: Optional[str] = None

    @property
    def total_hallazgos(self) -> int:
        return len(self.hallazgos)

    @property
    def nivel_salud(self) -> str:
        if self.porcentaje_datos_confiables >= 99.5:
            return "EXCELENTE"
        elif self.porcentaje_datos_confiables >= 98:
            return "BUENO"
        elif self.porcentaje_datos_confiables >= 95:
            return "REGULAR"
        elif self.porcentaje_datos_confiables >= 90:
            return "CRÍTICO"
        else:
            return "EMERGENCIA"


# ============================================================
# MOTOR DE AUDITORÍA DE CALIDAD DE DATOS
# ============================================================

class AuditoriaCalidadDatos:
    """
    Auditoría de calidad de datos sobre Odoo – solo lectura.

    Triple validación:
    1. Estado vs. Vínculo  → detecta Procesos Huérfanos
    2. Tiempo de Vida (SLA) → detecta registros Zombi sin actividad
    3. Incertidumbre         → calcula % de datos confiables vs basura
    """

    # Modelos que típicamente tienen operating_unit_id (OCA)
    _MODELOS_CON_OU = {
        'account.move', 'sale.order', 'purchase.order',
        'stock.picking', 'account.payment', 'crm.lead',
    }

    # SLA por defecto (días máx. sin actividad antes de ser Zombi)
    SLA_DEFAULTS = {
        'account.move': 30,         # Facturas abiertas > 30 días
        'sale.order': 15,           # Cotizaciones draft > 15 días
        'purchase.order': 15,       # OC en draft > 15 días
        'crm.lead': 30,             # Oportunidades sin actividad > 30 días
        'stock.picking': 7,         # Transferencias pendientes > 7 días
        'account.payment': 15,      # Pagos draft > 15 días
        'helpdesk.ticket': 5,       # Tickets sin atender > 5 días
    }

    def __init__(self, odoo=None, sla_dias: Optional[Dict[str, int]] = None):
        self.odoo = odoo
        self.sla = {**self.SLA_DEFAULTS, **(sla_dias or {})}
        self.hallazgos: List[HallazgoCalidad] = []
        self._ou_disponible: Optional[bool] = None

    def set_conector(self, odoo):
        self.odoo = odoo

    # ----------------------------------------------------------------
    # Helpers de trazabilidad (Empresa / Unidad Operativa / Usuario)
    # ----------------------------------------------------------------

    def _ou_check(self) -> bool:
        """Detecta si operating_unit_id está disponible en Odoo."""
        if self._ou_disponible is None:
            try:
                info = self.odoo.obtener_campos('account.move', ['operating_unit_id'])
                self._ou_disponible = 'operating_unit_id' in info
            except Exception:
                self._ou_disponible = False
        return self._ou_disponible

    def _enriquecer_hallazgos(self):
        """Enriquece hallazgos con empresa, unidad operativa y usuario creador (batch por modelo)."""
        if not self._hay_conexion():
            return
        por_modelo = defaultdict(list)
        for h in self.hallazgos:
            por_modelo[h.modelo_odoo].append(h)
        for modelo, hallazgos_modelo in por_modelo.items():
            ids = list({h.registro_id for h in hallazgos_modelo if h.registro_id})
            if not ids:
                continue
            try:
                campos = ['company_id', 'create_uid']
                if self._ou_check() and modelo in self._MODELOS_CON_OU:
                    campos.append('operating_unit_id')
                registros = self.odoo.search_read(modelo, [('id', 'in', ids)], campos=campos, limite=len(ids))
                mapa = {r['id']: r for r in registros}
                for h in hallazgos_modelo:
                    r = mapa.get(h.registro_id, {})
                    h.empresa = self._nombre_m2o(r.get('company_id'))
                    h.usuario_creador = self._nombre_m2o(r.get('create_uid'))
                    if r.get('operating_unit_id'):
                        h.unidad_operativa = self._nombre_m2o(r.get('operating_unit_id'))
            except Exception as e:
                print(f"  ⚠ Enriquecimiento {modelo}: {e}")

    # ================================================================
    # EJECUCIÓN PRINCIPAL
    # ================================================================

    def ejecutar_auditoria_completa(self) -> ResultadoCalidadDatos:
        """Ejecuta las tres fases de validación y genera Excel."""
        print("Iniciando Auditoría de Calidad de Datos…")
        self.hallazgos = []
        total_analizados = 0

        # ---- Fase 1: Estado vs Vínculo (Huérfanos) ----
        n1 = self._fase_estado_vs_vinculo()
        total_analizados += n1

        # ---- Fase 2: Tiempo de Vida / SLA (Zombis) ----
        n2 = self._fase_sla_zombis()
        total_analizados += n2

        # ---- Fase 3: Datos Incompletos / Inconsistentes ----
        n3 = self._fase_datos_incompletos()
        total_analizados += n3

        # ---- Enriquecer con Empresa / Unidad Operativa / Usuario ----
        self._enriquecer_hallazgos()

        total_analizados = max(total_analizados, 1)  # evitar /0

        # Calcular métricas de incertidumbre
        total_problemas = len(self.hallazgos)
        pct_basura = min(100.0, (total_problemas / total_analizados) * 100)
        pct_confiable = 100.0 - pct_basura
        indice_incertidumbre = pct_basura  # simplificado

        # Resúmenes
        por_cat = defaultdict(int)
        por_sev = defaultdict(int)
        por_modelo = defaultdict(int)
        por_empresa = defaultdict(int)
        por_ou = defaultdict(int)
        por_usuario = defaultdict(int)
        for h in self.hallazgos:
            por_cat[h.categoria] += 1
            por_sev[h.severidad] += 1
            por_modelo[h.modelo_odoo] += 1
            if h.empresa and h.empresa != 'N/A':
                por_empresa[h.empresa] += 1
            if h.unidad_operativa and h.unidad_operativa != 'N/A':
                por_ou[h.unidad_operativa] += 1
            if h.usuario_creador and h.usuario_creador != 'N/A':
                por_usuario[h.usuario_creador] += 1
        top_usuarios = sorted(por_usuario.items(), key=lambda x: -x[1])[:10]

        resultado = ResultadoCalidadDatos(
            fecha_ejecucion=datetime.now(),
            total_registros_analizados=total_analizados,
            hallazgos=self.hallazgos,
            resumen_por_categoria=dict(por_cat),
            resumen_por_severidad=dict(por_sev),
            resumen_por_modelo=dict(por_modelo),
            porcentaje_datos_confiables=round(pct_confiable, 2),
            porcentaje_datos_basura=round(pct_basura, 2),
            indice_incertidumbre=round(indice_incertidumbre, 2),
            resumen_por_empresa=dict(por_empresa),
            resumen_por_unidad_operativa=dict(por_ou),
            top_usuarios_problemas=top_usuarios,
        )

        # Generar Excel
        ruta = self._generar_excel(resultado)
        resultado.ruta_excel = ruta

        print(f"Auditoría finalizada: {total_problemas} hallazgos en {total_analizados} registros")
        return resultado

    # ================================================================
    # FASE 1 – Estado vs Vínculo (Procesos Huérfanos)
    # ================================================================

    def _fase_estado_vs_vinculo(self) -> int:
        """Detecta incoherencias entre estado y vínculos."""
        if not self._hay_conexion():
            return 0

        analizados = 0

        # 1A. Facturas ABIERTAS con pago vinculado
        analizados += self._validar_facturas_con_pago_huerfano()

        # 1B. Órdenes de venta confirmadas sin factura
        analizados += self._validar_ventas_sin_factura()

        # 1C. Transferencias de stock "done" sin origen
        analizados += self._validar_pickings_sin_origen()

        # 1D. Pagos publicados sin factura asociada
        analizados += self._validar_pagos_sin_factura()

        # 1E. Compras confirmadas sin recepción
        analizados += self._validar_compras_sin_recepcion()

        return analizados

    def _validar_facturas_con_pago_huerfano(self) -> int:
        """Factura abierta/draft PERO con pago vinculado → Proceso Incompleto."""
        try:
            # Move acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: TODAS las facturas no canceladas
            total_universo = self.odoo.contar('account.move', [
                ('move_type', 'in', ['out_invoice', 'in_invoice']),
                ('state', '!=', 'cancel'),
            ])
            # Subset problemático
            facturas = self.odoo.search_read('account.move', 
                [
                    ('move_type', 'in', ['out_invoice', 'in_invoice']),
                    ('state', '!=', 'cancel'),
                    ('payment_state', 'in', ['paid', 'in_payment', 'partial']),
                    ('amount_residual', '>', 0),
                ],
                ['name', 'partner_id', 'invoice_date', 'amount_total',
                 'amount_residual', 'state', 'payment_state', 'move_type'],
                limit=500
            )
            for f in facturas:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='incompleto',
                    severidad='critico',
                    modelo_odoo='account.move',
                    registro_id=f.get('id', 0),
                    registro_nombre=f.get('name', ''),
                    descripcion='Factura con pago vinculado pero saldo pendiente',
                    detalle=(
                        f"Estado pago: {f.get('payment_state')}, "
                        f"Residual: ${f.get('amount_residual', 0):,.2f} de ${f.get('amount_total', 0):,.2f}"
                    ),
                    accion_sugerida='Conciliar pago o revisar la aplicación del pago',
                    datos_extra={
                        'Cliente/Proveedor': self._nombre_partner(f.get('partner_id')),
                        'Fecha': str(f.get('invoice_date', '')),
                        'Tipo': 'Cliente' if 'out_' in str(f.get('move_type', '')) else 'Proveedor',
                    }
                ))
            return max(total_universo, len(facturas))
        except Exception as e:
            print(f"Facturas con pago incompleto: {e}")
            return 0

    def _validar_ventas_sin_factura(self) -> int:
        """OV confirmada (sale/done) con invoice_status ≠ invoiced."""
        try:
            # SO acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todas las ventas confirmadas
            total_universo = self.odoo.contar('sale.order', [('state', 'in', ['sale', 'done'])])
            ordenes = self.odoo.search_read('sale.order', 
                [
                    ('state', 'in', ['sale', 'done']),
                    ('invoice_status', '!=', 'invoiced'),
                ],
                ['name', 'partner_id', 'date_order', 'amount_total', 'invoice_status', 'state'],
                limit=500
            )
            for o in ordenes:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='incompleto',
                    severidad='alto',
                    modelo_odoo='sale.order',
                    registro_id=o.get('id', 0),
                    registro_nombre=o.get('name', ''),
                    descripcion='Venta confirmada sin facturar completamente',
                    detalle=f"Estado factura: {o.get('invoice_status')}, Total: ${o.get('amount_total', 0):,.2f}",
                    accion_sugerida='Generar factura o verificar entregas parciales',
                    datos_extra={
                        'Cliente': self._nombre_partner(o.get('partner_id')),
                        'Fecha': str(o.get('date_order', ''))[:10],
                    }
                ))
            return max(total_universo, len(ordenes))
        except Exception as e:
            print(f"Ventas sin factura: {e}")
            return 0

    def _validar_pickings_sin_origen(self) -> int:
        """Transferencias completadas sin documento de origen."""
        try:
            # Pick acceso encapsulado via ConectorOdoo (ARQ-003)
            fecha_90 = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            # Universo: todas las transferencias completadas en los últimos 90 días
            total_universo = self.odoo.contar('stock.picking', [
                ('state', '=', 'done'),
                ('date_done', '>=', fecha_90),
            ])
            pickings = self.odoo.search_read('stock.picking', 
                [
                    ('state', '=', 'done'),
                    ('origin', '=', False),
                    ('date_done', '>=', fecha_90),
                ],
                ['name', 'partner_id', 'date_done', 'picking_type_id', 'origin'],
                limit=300
            )
            for p in pickings:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='incompleto',
                    severidad='medio',
                    modelo_odoo='stock.picking',
                    registro_id=p.get('id', 0),
                    registro_nombre=p.get('name', ''),
                    descripcion='Transferencia completada sin documento de origen',
                    detalle=f"Tipo: {self._nombre_m2o(p.get('picking_type_id'))}, Fecha: {str(p.get('date_done', ''))[:10]}",
                    accion_sugerida='Verificar que la transferencia corresponde a una OV/OC',
                    datos_extra={
                        'Contacto': self._nombre_partner(p.get('partner_id')),
                    }
                ))
            return max(total_universo, len(pickings))
        except Exception as e:
            print(f" Pickings sin origen: {e}")
            return 0

    def _validar_pagos_sin_factura(self) -> int:
        """Pagos publicados que no están conciliados con ninguna factura."""
        try:
            # Payment acceso encapsulado via ConectorOdoo (ARQ-003)
            fecha_90 = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            # Universo: todos los pagos publicados en 90 días
            total_universo = self.odoo.contar('account.payment', [
                ('state', '=', 'posted'),
                ('date', '>=', fecha_90),
            ])
            pagos = self.odoo.search_read('account.payment', 
                [
                    ('state', '=', 'posted'),
                    ('is_reconciled', '=', False),
                    ('date', '>=', fecha_90),
                ],
                ['name', 'partner_id', 'amount', 'date', 'payment_type'],
                limit=500
            )
            for p in pagos:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='incompleto',
                    severidad='critico',
                    modelo_odoo='account.payment',
                    registro_id=p.get('id', 0),
                    registro_nombre=p.get('name', ''),
                    descripcion='Pago publicado sin conciliar con factura',
                    detalle=f"Monto: ${p.get('amount', 0):,.2f}, Tipo: {p.get('payment_type')}",
                    accion_sugerida='Conciliar con factura correspondiente o investigar origen',
                    datos_extra={
                        'Contacto': self._nombre_partner(p.get('partner_id')),
                        'Fecha': str(p.get('date', '')),
                    }
                ))
            return max(total_universo, len(pagos))
        except Exception as e:
            print(f"  Pagos sin factura: {e}")
            return 0

    def _validar_compras_sin_recepcion(self) -> int:
        """Compras confirmadas sin recepciones asociadas."""
        try:
            # PO acceso encapsulado via ConectorOdoo (ARQ-003)
            fecha_90 = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            # Universo: todas las compras confirmadas en 90 días
            total_universo = self.odoo.contar('purchase.order', [
                ('state', '=', 'purchase'),
                ('date_order', '>=', fecha_90),
            ])
            compras = self.odoo.search_read('purchase.order', 
                [
                    ('state', '=', 'purchase'),
                    ('receipt_status', '!=', 'full'),
                    ('date_order', '>=', fecha_90),
                ],
                ['name', 'partner_id', 'date_order', 'amount_total', 'receipt_status'],
                limit=300
            )
            for c in compras:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='incompleto',
                    severidad='alto',
                    modelo_odoo='purchase.order',
                    registro_id=c.get('id', 0),
                    registro_nombre=c.get('name', ''),
                    descripcion='Compra confirmada sin recepción completa',
                    detalle=f"Estado recepción: {c.get('receipt_status')}, Total: ${c.get('amount_total', 0):,.2f}",
                    accion_sugerida='Verificar con almacén si se recibió mercancía',
                    datos_extra={
                        'Proveedor': self._nombre_partner(c.get('partner_id')),
                        'Fecha': str(c.get('date_order', ''))[:10],
                    }
                ))
            return max(total_universo, len(compras))
        except Exception as e:
            print(f"  Compras sin recepción: {e}")
            return 0

    # ================================================================
    # FASE 2 – Tiempo de Vida / SLA (Registros Zombi)
    # ================================================================

    def _fase_sla_zombis(self) -> int:
        """Detecta registros que exceden su SLA sin actividad."""
        if not self._hay_conexion():
            return 0

        analizados = 0
        hoy = datetime.now()

        # 2A. Facturas abiertas viejas
        analizados += self._zombis_facturas(hoy)

        # 2B. Cotizaciones draft abandonadas
        analizados += self._zombis_cotizaciones(hoy)

        # 2C. Compras draft abandonadas
        analizados += self._zombis_compras_draft(hoy)

        # 2D. Transferencias pendientes sin mover
        analizados += self._zombis_pickings(hoy)

        # 2E. Oportunidades CRM estancadas
        analizados += self._zombis_crm(hoy)

        return analizados

    def _zombis_facturas(self, hoy: datetime) -> int:
        """Facturas abiertas/draft que superan el SLA."""
        sla = self.sla.get('account.move', 30)
        fecha_limite = (hoy - timedelta(days=sla)).strftime('%Y-%m-%d')
        try:
            # Move acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todas las facturas en draft
            total_universo = self.odoo.contar('account.move', [
                ('move_type', 'in', ['out_invoice', 'in_invoice']),
                ('state', '=', 'draft'),
            ])
            facturas = self.odoo.search_read('account.move', 
                [
                    ('move_type', 'in', ['out_invoice', 'in_invoice']),
                    ('state', '=', 'draft'),
                    ('create_date', '<=', fecha_limite),
                ],
                ['name', 'partner_id', 'create_date', 'amount_total', 'move_type'],
                limit=500
            )
            for f in facturas:
                dias = self._dias_desde(f.get('create_date'))
                self.hallazgos.append(HallazgoCalidad(
                    categoria='zombi',
                    severidad='alto' if dias > sla * 2 else 'medio',
                    modelo_odoo='account.move',
                    registro_id=f.get('id', 0),
                    registro_nombre=f.get('name', ''),
                    descripcion=f'Factura en borrador desde hace {dias} días (SLA: {sla}d)',
                    detalle=f"Total: ${f.get('amount_total', 0):,.2f}, Creada: {str(f.get('create_date', ''))[:10]}",
                    accion_sugerida='Publicar, cancelar o eliminar la factura borrador',
                    datos_extra={
                        'Días Zombi': dias,
                        'SLA Esperado': sla,
                        'Contacto': self._nombre_partner(f.get('partner_id')),
                    }
                ))
            return max(total_universo, len(facturas))
        except Exception as e:
            print(f"  Zombis facturas: {e}")
            return 0

    def _zombis_cotizaciones(self, hoy: datetime) -> int:
        """Cotizaciones draft sin convertir en venta."""
        sla = self.sla.get('sale.order', 15)
        fecha_limite = (hoy - timedelta(days=sla)).strftime('%Y-%m-%d')
        try:
            # SO acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todas las cotizaciones en draft
            total_universo = self.odoo.contar('sale.order', [('state', '=', 'draft')])
            cotizaciones = self.odoo.search_read('sale.order', 
                [
                    ('state', '=', 'draft'),
                    ('create_date', '<=', fecha_limite),
                ],
                ['name', 'partner_id', 'create_date', 'amount_total'],
                limit=500
            )
            for c in cotizaciones:
                dias = self._dias_desde(c.get('create_date'))
                self.hallazgos.append(HallazgoCalidad(
                    categoria='zombi',
                    severidad='medio',
                    modelo_odoo='sale.order',
                    registro_id=c.get('id', 0),
                    registro_nombre=c.get('name', ''),
                    descripcion=f'Cotización abandonada desde hace {dias} días (SLA: {sla}d)',
                    detalle=f"Total: ${c.get('amount_total', 0):,.2f}",
                    accion_sugerida='Contactar al cliente o cancelar la cotización',
                    datos_extra={
                        'Días Zombi': dias,
                        'Cliente': self._nombre_partner(c.get('partner_id')),
                    }
                ))
            return max(total_universo, len(cotizaciones))
        except Exception as e:
            print(f" Zombis cotizaciones: {e}")
            return 0

    def _zombis_compras_draft(self, hoy: datetime) -> int:
        """OC en draft por mucho tiempo."""
        sla = self.sla.get('purchase.order', 15)
        fecha_limite = (hoy - timedelta(days=sla)).strftime('%Y-%m-%d')
        try:
            # PO acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todas las OC en draft
            total_universo = self.odoo.contar('purchase.order', [('state', '=', 'draft')])
            compras = self.odoo.search_read('purchase.order', 
                [
                    ('state', '=', 'draft'),
                    ('create_date', '<=', fecha_limite),
                ],
                ['name', 'partner_id', 'create_date', 'amount_total'],
                limit=300
            )
            for c in compras:
                dias = self._dias_desde(c.get('create_date'))
                self.hallazgos.append(HallazgoCalidad(
                    categoria='zombi',
                    severidad='medio',
                    modelo_odoo='purchase.order',
                    registro_id=c.get('id', 0),
                    registro_nombre=c.get('name', ''),
                    descripcion=f'Orden de compra draft desde hace {dias} días (SLA: {sla}d)',
                    detalle=f"Total: ${c.get('amount_total', 0):,.2f}",
                    accion_sugerida='Confirmar, cancelar o eliminar la OC',
                    datos_extra={
                        'Días Zombi': dias,
                        'Proveedor': self._nombre_partner(c.get('partner_id')),
                    }
                ))
            return max(total_universo, len(compras))
        except Exception as e:
            print(f"  Zombis compras: {e}")
            return 0

    def _zombis_pickings(self, hoy: datetime) -> int:
        """Transferencias pendientes por mucho tiempo."""
        sla = self.sla.get('stock.picking', 7)
        fecha_limite = (hoy - timedelta(days=sla)).strftime('%Y-%m-%d')
        try:
            # Pick acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todas las transferencias pendientes
            total_universo = self.odoo.contar('stock.picking', [
                ('state', 'in', ['confirmed', 'assigned', 'waiting']),
            ])
            pickings = self.odoo.search_read('stock.picking', 
                [
                    ('state', 'in', ['confirmed', 'assigned', 'waiting']),
                    ('scheduled_date', '<=', fecha_limite),
                ],
                ['name', 'partner_id', 'scheduled_date', 'picking_type_id', 'origin', 'state'],
                limit=300
            )
            for p in pickings:
                dias = self._dias_desde(p.get('scheduled_date'))
                self.hallazgos.append(HallazgoCalidad(
                    categoria='zombi',
                    severidad='alto' if dias > sla * 3 else 'medio',
                    modelo_odoo='stock.picking',
                    registro_id=p.get('id', 0),
                    registro_nombre=p.get('name', ''),
                    descripcion=f'Transferencia pendiente desde hace {dias} días (SLA: {sla}d)',
                    detalle=f"Estado: {p.get('state')}, Origen: {p.get('origin', 'N/A')}",
                    accion_sugerida='Procesar la transferencia o cancelar si ya no aplica',
                    datos_extra={
                        'Días Zombi': dias,
                        'Tipo': self._nombre_m2o(p.get('picking_type_id')),
                    }
                ))
            return max(total_universo, len(pickings))
        except Exception as e:
            print(f"  Zombis pickings: {e}")
            return 0

    def _zombis_crm(self, hoy: datetime) -> int:
        """Oportunidades CRM estancadas sin actividad."""
        sla = self.sla.get('crm.lead', 30)
        fecha_limite = (hoy - timedelta(days=sla)).strftime('%Y-%m-%d')
        try:
            # Lead acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todas las oportunidades activas no ganadas
            total_universo = self.odoo.contar('crm.lead', [
                ('active', '=', True),
                ('probability', '<', 100),
            ])
            leads = self.odoo.search_read('crm.lead', 
                [
                    ('active', '=', True),
                    ('probability', '<', 100),
                    ('write_date', '<=', fecha_limite),
                ],
                ['name', 'partner_id', 'write_date', 'expected_revenue',
                 'stage_id', 'user_id', 'probability'],
                limit=300
            )
            for l in leads:
                dias = self._dias_desde(l.get('write_date'))
                self.hallazgos.append(HallazgoCalidad(
                    categoria='zombi',
                    severidad='medio',
                    modelo_odoo='crm.lead',
                    registro_id=l.get('id', 0),
                    registro_nombre=l.get('name', ''),
                    descripcion=f'Oportunidad CRM sin actividad desde hace {dias} días',
                    detalle=f"Etapa: {self._nombre_m2o(l.get('stage_id'))}, Prob: {l.get('probability', 0)}%",
                    accion_sugerida='Dar seguimiento al prospecto o archivar la oportunidad',
                    datos_extra={
                        'Días Zombi': dias,
                        'Ingreso Esperado': f"${l.get('expected_revenue', 0):,.2f}",
                        'Vendedor': self._nombre_m2o(l.get('user_id')),
                    }
                ))
            return max(total_universo, len(leads))
        except Exception as e:
            print(f"  Zombis CRM: {e}")
            return 0

    # ================================================================
    # FASE 3 – Datos Incompletos / Inconsistentes
    # ================================================================

    def _fase_datos_incompletos(self) -> int:
        """Registros con campos clave vacíos o inconsistentes."""
        if not self._hay_conexion():
            return 0

        analizados = 0

        # 3A. Clientes sin datos de contacto
        analizados += self._clientes_sin_contacto()

        # 3B. Productos sin precio o con precio 0
        analizados += self._productos_sin_precio()

        # 3C. Facturas publicadas con total 0
        analizados += self._facturas_total_cero()

        # 3D. Ventas con líneas sin producto
        analizados += self._ventas_lineas_sin_producto()

        # 3E. Movimientos de inventario con cantidad 0
        analizados += self._stock_cantidad_cero()

        return analizados

    def _clientes_sin_contacto(self) -> int:
        """Clientes activos sin email ni teléfono."""
        try:
            # Partner acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todos los clientes activos
            total_universo = self.odoo.contar('res.partner', [
                ('customer_rank', '>', 0),
                ('active', '=', True),
            ])
            clientes = self.odoo.search_read('res.partner', 
                [
                    ('customer_rank', '>', 0),
                    ('active', '=', True),
                    ('email', '=', False),
                    ('phone', '=', False),
                    ('mobile', '=', False),
                ],
                ['name', 'customer_rank', 'create_date'],
                limit=500
            )
            for c in clientes:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='incompleto',
                    severidad='medio',
                    modelo_odoo='res.partner',
                    registro_id=c.get('id', 0),
                    registro_nombre=c.get('name', ''),
                    descripcion='Cliente sin datos de contacto (email, teléfono, celular)',
                    detalle=f"Rank: {c.get('customer_rank')}, Creado: {str(c.get('create_date', ''))[:10]}",
                    accion_sugerida='Completar información de contacto del cliente',
                ))
            return max(total_universo, len(clientes))
        except Exception as e:
            print(f"  Clientes sin contacto: {e}")
            return 0

    def _productos_sin_precio(self) -> int:
        """Productos activos con precio de venta 0."""
        try:
            # Prod acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todos los productos vendibles activos
            total_universo = self.odoo.contar('product.template', [
                ('active', '=', True),
                ('sale_ok', '=', True),
            ])
            productos = self.odoo.search_read('product.template', 
                [
                    ('active', '=', True),
                    ('sale_ok', '=', True),
                    ('list_price', '<=', 0),
                ],
                ['name', 'default_code', 'list_price', 'type'],
                limit=500
            )
            for p in productos:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='incompleto',
                    severidad='alto',
                    modelo_odoo='product.template',
                    registro_id=p.get('id', 0),
                    registro_nombre=p.get('name', ''),
                    descripcion='Producto vendible con precio $0 o negativo',
                    detalle=f"Precio: ${p.get('list_price', 0):,.2f}, Código: {p.get('default_code', 'N/A')}",
                    accion_sugerida='Asignar precio de venta correcto al producto',
                ))
            return max(total_universo, len(productos))
        except Exception as e:
            print(f"  Productos sin precio: {e}")
            return 0

    def _facturas_total_cero(self) -> int:
        """Facturas publicadas con total = 0."""
        try:
            # Move acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todas las facturas publicadas
            total_universo = self.odoo.contar('account.move', [
                ('move_type', 'in', ['out_invoice', 'in_invoice']),
                ('state', '=', 'posted'),
            ])
            facturas = self.odoo.search_read('account.move', 
                [
                    ('move_type', 'in', ['out_invoice', 'in_invoice']),
                    ('state', '=', 'posted'),
                    ('amount_total', '=', 0),
                ],
                ['name', 'partner_id', 'invoice_date', 'move_type'],
                limit=300
            )
            for f in facturas:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='inconsistente',
                    severidad='critico',
                    modelo_odoo='account.move',
                    registro_id=f.get('id', 0),
                    registro_nombre=f.get('name', ''),
                    descripcion='Factura publicada con total $0',
                    detalle=f"Fecha: {f.get('invoice_date', 'N/A')}, Tipo: {'Cliente' if 'out_' in str(f.get('move_type', '')) else 'Proveedor'}",
                    accion_sugerida='Revisar líneas de factura o cancelar',
                    datos_extra={
                        'Contacto': self._nombre_partner(f.get('partner_id')),
                    }
                ))
            return max(total_universo, len(facturas))
        except Exception as e:
            print(f"  Facturas total cero: {e}")
            return 0

    def _ventas_lineas_sin_producto(self) -> int:
        """Líneas de venta sin producto asignado."""
        try:
            # SOL acceso encapsulado via ConectorOdoo (ARQ-003)
            # Universo: todas las líneas de ventas confirmadas
            total_universo = self.odoo.contar('sale.order.line', [
                ('order_id.state', 'in', ['sale', 'done']),
            ])
            lineas = self.odoo.search_read('sale.order.line', 
                [
                    ('product_id', '=', False),
                    ('order_id.state', 'in', ['sale', 'done']),
                ],
                ['order_id', 'name', 'price_unit', 'product_uom_qty'],
                limit=300
            )
            for l in lineas:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='incompleto',
                    severidad='medio',
                    modelo_odoo='sale.order.line',
                    registro_id=l.get('id', 0),
                    registro_nombre=self._nombre_m2o(l.get('order_id')),
                    descripcion='Línea de venta confirmada sin producto asignado',
                    detalle=f"Desc: {(l.get('name', '') or '')[:60]}, Precio: ${l.get('price_unit', 0):,.2f}",
                    accion_sugerida='Asignar producto a la línea de venta',
                ))
            return max(total_universo, len(lineas))
        except Exception as e:
            print(f"  Líneas sin producto: {e}")
            return 0

    def _stock_cantidad_cero(self) -> int:
        """Movimientos de stock con cantidad 0 completados."""
        try:
            # SMove acceso encapsulado via ConectorOdoo (ARQ-003)
            fecha_90 = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
            # Universo: todos los movimientos de stock completados en 90 días
            total_universo = self.odoo.contar('stock.move', [
                ('state', '=', 'done'),
                ('date', '>=', fecha_90),
            ])
            movimientos = self.odoo.search_read('stock.move',
                [
                    ('state', '=', 'done'),
                    ('quantity_done', '=', 0),
                    ('product_uom_qty', '>', 0),
                    ('date', '>=', fecha_90),
                ],
                ['reference', 'product_id', 'product_uom_qty', 'quantity_done', 'date'],
                limit=300
            )
            for m in movimientos:
                self.hallazgos.append(HallazgoCalidad(
                    categoria='inconsistente',
                    severidad='alto',
                    modelo_odoo='stock.move',
                    registro_id=m.get('id', 0),
                    registro_nombre=m.get('reference', ''),
                    descripcion='Movimiento completado con cantidad hecha = 0',
                    detalle=f"Demanda: {m.get('product_uom_qty', 0)}, Hecho: 0, Producto: {self._nombre_m2o(m.get('product_id'))}",
                    accion_sugerida='Verificar si el movimiento debería revertirse',
                ))
            return max(total_universo, len(movimientos))
        except Exception as e:
            print(f"  Stock cantidad cero: {e}")
            return 0

    # ================================================================
    # GENERACIÓN DE EXCEL
    # ================================================================

    def _generar_excel(self, resultado: ResultadoCalidadDatos) -> Optional[str]:
        """Genera un Excel profesional con todos los hallazgos."""
        if not PANDAS_DISPONIBLE:
            print("  pandas no disponible, no se puede generar Excel")
            return None

        if not resultado.hallazgos:
            print("  ℹ Sin hallazgos, no se genera Excel")
            return None

        # Preparar directorio
        carpeta = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Reportes_Bot')
        os.makedirs(carpeta, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ruta = os.path.join(carpeta, f'Auditoria_Calidad_Datos_{timestamp}.xlsx')

        # Crear DataFrame principal
        filas = [h.to_dict() for h in resultado.hallazgos]
        df_hallazgos = pd.DataFrame(filas)

        # Resumen
        df_resumen = pd.DataFrame([
            {'Métrica': 'Fecha de Ejecución', 'Valor': resultado.fecha_ejecucion.strftime('%Y-%m-%d %H:%M')},
            {'Métrica': 'Total Registros Analizados', 'Valor': resultado.total_registros_analizados},
            {'Métrica': 'Total Hallazgos', 'Valor': resultado.total_hallazgos},
            {'Métrica': '% Datos Confiables', 'Valor': f"{resultado.porcentaje_datos_confiables:.1f}%"},
            {'Métrica': '% Datos con Problemas', 'Valor': f"{resultado.porcentaje_datos_basura:.1f}%"},
            {'Métrica': 'Índice de Incertidumbre', 'Valor': f"{resultado.indice_incertidumbre:.1f}%"},
            {'Métrica': 'Nivel de Salud', 'Valor': resultado.nivel_salud},
        ])

        # Por categoría
        df_categorias = pd.DataFrame([
            {'Categoría': k, 'Cantidad': v}
            for k, v in resultado.resumen_por_categoria.items()
        ])

        # Por severidad
        df_severidad = pd.DataFrame([
            {'Severidad': k, 'Cantidad': v}
            for k, v in resultado.resumen_por_severidad.items()
        ])

        # Por modelo
        df_modelos = pd.DataFrame([
            {'Modelo Odoo': k, 'Cantidad': v}
            for k, v in resultado.resumen_por_modelo.items()
        ])

        # Por empresa
        total_h = max(resultado.total_hallazgos, 1)
        df_empresas = pd.DataFrame([
            {'Empresa': k, 'Hallazgos': v, '% del Total': f"{(v/total_h)*100:.1f}%"}
            for k, v in sorted(resultado.resumen_por_empresa.items(), key=lambda x: -x[1])
        ]) if resultado.resumen_por_empresa else pd.DataFrame()

        # Por unidad operativa
        df_ou = pd.DataFrame([
            {'Unidad Operativa': k, 'Hallazgos': v, '% del Total': f"{(v/total_h)*100:.1f}%"}
            for k, v in sorted(resultado.resumen_por_unidad_operativa.items(), key=lambda x: -x[1])
        ]) if resultado.resumen_por_unidad_operativa else pd.DataFrame()

        # Top usuarios
        df_usuarios = pd.DataFrame([
            {'#': i+1, 'Usuario': u, 'Hallazgos': c, '% del Total': f"{(c/total_h)*100:.1f}%"}
            for i, (u, c) in enumerate(resultado.top_usuarios_problemas)
        ]) if resultado.top_usuarios_problemas else pd.DataFrame()

        try:
            with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
                df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
                df_hallazgos.to_excel(writer, sheet_name='Hallazgos', index=False)
                df_categorias.to_excel(writer, sheet_name='Por Categoría', index=False)
                df_severidad.to_excel(writer, sheet_name='Por Severidad', index=False)
                df_modelos.to_excel(writer, sheet_name='Por Modelo', index=False)
                if not df_empresas.empty:
                    df_empresas.to_excel(writer, sheet_name='Por Empresa', index=False)
                if not df_ou.empty:
                    df_ou.to_excel(writer, sheet_name='Por Unidad Operativa', index=False)
                if not df_usuarios.empty:
                    df_usuarios.to_excel(writer, sheet_name='Top Usuarios', index=False)

                # Estilizar si openpyxl disponible
                if OPENPYXL_DISPONIBLE:
                    self._estilizar_excel(writer)

            print(f"  📊 Excel generado: {ruta}")
            return ruta
        except Exception as e:
            logger.error(f"  ⚠ Error generando Excel: {e}")
            return None

    def _estilizar_excel(self, writer):
        """Aplica estilos profesionales al Excel."""
        wb = writer.book

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        critico_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        alto_fill = PatternFill(start_color='FFA726', end_color='FFA726', fill_type='solid')
        medio_fill = PatternFill(start_color='FFEE58', end_color='FFEE58', fill_type='solid')
        bajo_fill = PatternFill(start_color='66BB6A', end_color='66BB6A', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for ws in wb.worksheets:
            # Cabeceras
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Ajustar anchos
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            # Colorear severidad en hoja Hallazgos
            if ws.title == 'Hallazgos':
                sev_col = None
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == 'Severidad':
                        sev_col = idx
                        break
                if sev_col:
                    fill_map = {
                        'critico': critico_fill,
                        'alto': alto_fill,
                        'medio': medio_fill,
                        'bajo': bajo_fill,
                    }
                    for row in ws.iter_rows(min_row=2):
                        cell_sev = row[sev_col - 1]
                        fill = fill_map.get(str(cell_sev.value).lower())
                        if fill:
                            cell_sev.fill = fill

    # ================================================================
    # FORMATEO MARKDOWN PARA CHAT
    # ================================================================

    def formatear_resultado_markdown(self, resultado: ResultadoCalidadDatos) -> str:
        """Formatea el resultado para mostrar en el chatbot."""
        emoji_salud = {
            'EXCELENTE': '🟢', 'BUENO': '🟡', 'REGULAR': '🟠',
            'CRÍTICO': '🔴', 'EMERGENCIA': '🚨'
        }
        emoji = emoji_salud.get(resultado.nivel_salud, '⚪')

        md = f"""## 🔍 Auditoría de Calidad de Datos

### {emoji} Nivel de Salud: **{resultado.nivel_salud}**

| Métrica | Valor |
|---------|-------|
| Registros Analizados | **{resultado.total_registros_analizados:,}** |
| Hallazgos Encontrados | **{resultado.total_hallazgos:,}** |
| Datos Confiables | **{resultado.porcentaje_datos_confiables:.1f}%** |
| Datos con Problemas | **{resultado.porcentaje_datos_basura:.1f}%** |
| Índice de Incertidumbre | **{resultado.indice_incertidumbre:.1f}%** |

> ⚠️ **Ojo:** El **{resultado.porcentaje_datos_basura:.1f}%** de tus datos presentan inconsistencias. El análisis es confiable al **{resultado.porcentaje_datos_confiables:.1f}%**.

---

### 📊 Hallazgos por Categoría

| Categoría | Cantidad | Descripción |
|-----------|----------|-------------|
"""
        desc_cat = {
            'huerfano': '🔗 Procesos desvinculados (Estado vs Vínculo)',
            'zombi': '🧟 Registros sin actividad (exceden SLA)',
            'incompleto': '📝 Datos faltantes en campos clave',
            'inconsistente': '⚠️ Datos contradictorios o imposibles',
        }
        for cat, cant in sorted(resultado.resumen_por_categoria.items(), key=lambda x: -x[1]):
            md += f"| **{cat.capitalize()}** | {cant} | {desc_cat.get(cat, '')} |\n"

        md += "\n### 🚦 Por Severidad\n\n"
        emoji_sev = {'critico': '🔴', 'alto': '🟠', 'medio': '🟡', 'bajo': '🟢'}
        for sev, cant in sorted(resultado.resumen_por_severidad.items(),
                                key=lambda x: ['critico', 'alto', 'medio', 'bajo'].index(x[0])
                                if x[0] in ['critico', 'alto', 'medio', 'bajo'] else 99):
            md += f"- {emoji_sev.get(sev, '⚪')} **{sev.capitalize()}**: {cant}\n"

        md += "\n### 📋 Por Modelo Odoo\n\n"
        for mod, cant in sorted(resultado.resumen_por_modelo.items(), key=lambda x: -x[1]):
            md += f"- `{mod}`: {cant} hallazgos\n"

        # ---- Por Empresa ----
        if resultado.resumen_por_empresa:
            total_h = max(resultado.total_hallazgos, 1)
            md += "\n### 🏢 Hallazgos por Empresa\n\n"
            md += "| Empresa | Hallazgos | % del Total |\n"
            md += "|---------|-----------|-------------|\n"
            for emp, cant in sorted(resultado.resumen_por_empresa.items(), key=lambda x: -x[1]):
                pct = (cant / total_h) * 100
                md += f"| **{emp}** | {cant} | {pct:.1f}% |\n"

        # ---- Por Unidad Operativa ----
        if resultado.resumen_por_unidad_operativa:
            total_h = max(resultado.total_hallazgos, 1)
            md += "\n### 🏭 Hallazgos por Unidad Operativa\n\n"
            md += "| Unidad Operativa | Hallazgos | % del Total |\n"
            md += "|------------------|-----------|-------------|\n"
            for ou, cant in sorted(resultado.resumen_por_unidad_operativa.items(), key=lambda x: -x[1]):
                pct = (cant / total_h) * 100
                md += f"| **{ou}** | {cant} | {pct:.1f}% |\n"

        # ---- Top 3 Usuarios con más problemas ----
        if resultado.top_usuarios_problemas:
            md += "\n### 👤 Top Usuarios con más Hallazgos\n\n"
            medallas = ['🥇', '🥈', '🥉']
            for i, (usuario, cant) in enumerate(resultado.top_usuarios_problemas[:3]):
                medalla = medallas[i] if i < 3 else f"{i+1}."
                pct = (cant / max(resultado.total_hallazgos, 1)) * 100
                md += f"{medalla} **{usuario}** — {cant} hallazgos ({pct:.1f}% del total)\n"

        # Top 10 hallazgos críticos
        criticos = [h for h in resultado.hallazgos if h.severidad == 'critico'][:10]
        if criticos:
            md += "\n### 🔴 Top Hallazgos Críticos\n\n"
            for i, h in enumerate(criticos, 1):
                md += f"**{i}. [{h.registro_nombre}]** – {h.descripcion}\n"
                empresa_tag = f" | 🏢 {h.empresa}" if h.empresa and h.empresa != 'N/A' else ''
                usuario_tag = f" | 👤 {h.usuario_creador}" if h.usuario_creador and h.usuario_creador != 'N/A' else ''
                md += f"   > {h.detalle}{empresa_tag}{usuario_tag}\n"
                md += f"   💡 *{h.accion_sugerida}*\n\n"

        if resultado.ruta_excel:
            md += f"\n---\n📎 **Excel generado:** `{resultado.ruta_excel}`\n"

        return md

    # ================================================================
    # UTILIDADES INTERNAS
    # ================================================================

    def _hay_conexion(self) -> bool:
        if not self.odoo or not self.odoo.conectado:
            print("  ⚠ Sin conexión a Odoo para auditoría de calidad")
            return False
        return True

    @staticmethod
    def _nombre_partner(campo) -> str:
        if isinstance(campo, (list, tuple)) and len(campo) > 1:
            return str(campo[1])
        return str(campo) if campo else 'N/A'

    @staticmethod
    def _nombre_m2o(campo) -> str:
        if isinstance(campo, (list, tuple)) and len(campo) > 1:
            return str(campo[1])
        return str(campo) if campo else 'N/A'

    @staticmethod
    def _dias_desde(fecha_str) -> int:
        if not fecha_str:
            return 0
        try:
            fecha = datetime.strptime(str(fecha_str)[:10], '%Y-%m-%d')
            return (datetime.now() - fecha).days
        except Exception:
            return 0
