# ============================================================
# GENERADOR DE REPORTES PRO - Excel, PDF y HTML
# ============================================================
# Genera reportes profesionales con gráficos y formato
# ============================================================

import pandas as pd
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from io import BytesIO

# Para Excel con formato
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.drawing.image import Image
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False
    print("openpyxl no instalado. Instalando: pip install openpyxl")

# Para gráficos
try:
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')  # Backend sin GUI
    MATPLOTLIB_DISPONIBLE = True
except ImportError:
    MATPLOTLIB_DISPONIBLE = False

# Para PDF
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.units import inch
    REPORTLAB_DISPONIBLE = True
except ImportError:
    REPORTLAB_DISPONIBLE = False


class GeneradorReportes:
    """
    Generador de reportes profesionales en múltiples formatos.
    """
    
    def __init__(self, directorio_salida: str = None):
        """
        Inicializa el generador de reportes.
        
        Args:
            directorio_salida: Directorio donde se guardarán los reportes
        """
        self.directorio = directorio_salida or os.path.join(
            os.path.dirname(__file__), '..', 'Reportes_Bot'
        )
        os.makedirs(self.directorio, exist_ok=True)
        
        # Colores corporativos
        self.colores = {
            'primario': '1F4E79',
            'secundario': '2E75B6',
            'acento': '00B050',
            'alerta': 'FF6600',
            'error': 'C00000',
            'fondo_header': '1F4E79',
            'fondo_alterno': 'D6DCE4',
        }
        
        # Estilos Excel
        self.estilos = self._crear_estilos_excel() if OPENPYXL_DISPONIBLE else {}
    
    def _crear_estilos_excel(self) -> Dict:
        """Crea estilos para Excel."""
        return {
            'titulo': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
            'subtitulo': Font(name='Calibri', size=14, bold=True, color=self.colores['primario']),
            'header': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'normal': Font(name='Calibri', size=10),
            'moneda': Font(name='Calibri', size=10),
            'numero': Font(name='Calibri', size=10),
            'fill_header': PatternFill(start_color=self.colores['fondo_header'], 
                                       end_color=self.colores['fondo_header'], 
                                       fill_type='solid'),
            'fill_alterno': PatternFill(start_color=self.colores['fondo_alterno'],
                                        end_color=self.colores['fondo_alterno'],
                                        fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center'),
        }
    
    def _generar_nombre_archivo(self, prefijo: str, extension: str) -> str:
        """Genera un nombre de archivo único con timestamp."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return os.path.join(self.directorio, f"{prefijo}_{timestamp}.{extension}")
    
    # ========================================
    # REPORTES EN EXCEL
    # ========================================
    
    def crear_excel_profesional(self, datos: Dict[str, pd.DataFrame], 
                                titulo: str = "Reporte",
                                incluir_graficos: bool = True) -> str:
        """
        Crea un archivo Excel profesional con múltiples hojas y gráficos.
        
        Args:
            datos: Diccionario {nombre_hoja: DataFrame}
            titulo: Título del reporte
            incluir_graficos: Si incluir gráficos automáticos
        
        Returns:
            Ruta del archivo creado
        """
        if not OPENPYXL_DISPONIBLE:
            # Fallback a pandas
            return self._crear_excel_simple(datos, titulo)
        
        wb = Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        for nombre_hoja, df in datos.items():
            if df.empty:
                continue
            
            ws = wb.create_sheet(title=nombre_hoja[:31])  # Excel limita a 31 chars
            
            # Agregar título
            ws.merge_cells('A1:F1')
            celda_titulo = ws['A1']
            celda_titulo.value = f"{titulo} - {nombre_hoja}"
            celda_titulo.font = self.estilos['titulo']
            celda_titulo.fill = self.estilos['fill_header']
            celda_titulo.alignment = self.estilos['center']
            ws.row_dimensions[1].height = 35
            
            # Agregar fecha de generación
            ws.merge_cells('A2:F2')
            ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(italic=True, size=9, color='666666')
            
            # Espacio
            fila_inicio = 4
            
            # Escribir encabezados
            for col_idx, columna in enumerate(df.columns, 1):
                celda = ws.cell(row=fila_inicio, column=col_idx)
                celda.value = str(columna).replace('_', ' ').title()
                celda.font = self.estilos['header']
                celda.fill = self.estilos['fill_header']
                celda.border = self.estilos['border']
                celda.alignment = self.estilos['center']
            
            # Escribir datos
            for row_idx, fila in enumerate(df.values, fila_inicio + 1):
                for col_idx, valor in enumerate(fila, 1):
                    celda = ws.cell(row=row_idx, column=col_idx)
                    
                    # Procesar valores especiales de Odoo
                    if isinstance(valor, (list, tuple)):
                        valor = valor[1] if len(valor) > 1 else str(valor)
                    
                    celda.value = valor
                    celda.font = self.estilos['normal']
                    celda.border = self.estilos['border']
                    
                    # Formato según tipo de dato
                    if isinstance(valor, (int, float)):
                        celda.alignment = self.estilos['right']
                        if 'total' in str(df.columns[col_idx-1]).lower() or \
                           'monto' in str(df.columns[col_idx-1]).lower() or \
                           'precio' in str(df.columns[col_idx-1]).lower():
                            celda.number_format = '$#,##0.00'
                    else:
                        celda.alignment = self.estilos['left']
                    
                    # Fila alterna
                    if row_idx % 2 == 0:
                        celda.fill = self.estilos['fill_alterno']
            
            # Ajustar ancho de columnas
            for col_idx, columna in enumerate(df.columns, 1):
                max_length = max(
                    len(str(columna)),
                    df[columna].astype(str).str.len().max() if not df[columna].empty else 0
                )
                ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)
            
            # Agregar gráfico si hay datos numéricos
            if incluir_graficos and len(df) > 1:
                self._agregar_grafico_excel(ws, df, fila_inicio, len(df))
        
        # Agregar hoja de resumen si hay múltiples hojas
        if len(datos) > 1:
            self._agregar_hoja_resumen(wb, datos)
        
        # Guardar
        archivo = self._generar_nombre_archivo(titulo.replace(' ', '_'), 'xlsx')
        wb.save(archivo)
        
        return archivo
    
    def _crear_excel_simple(self, datos: Dict[str, pd.DataFrame], titulo: str) -> str:
        """Crea Excel simple sin formato (fallback)."""
        archivo = self._generar_nombre_archivo(titulo.replace(' ', '_'), 'xlsx')
        
        with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
            for nombre, df in datos.items():
                if not df.empty:
                    df.to_excel(writer, sheet_name=nombre[:31], index=False)
        
        return archivo
    
    def _agregar_grafico_excel(self, ws, df: pd.DataFrame, fila_inicio: int, 
                                num_filas: int):
        """Agrega un gráfico a la hoja de Excel."""
        # Buscar columna numérica para el gráfico
        col_numerica = None
        for idx, col in enumerate(df.columns):
            if df[col].dtype in ['int64', 'float64']:
                if 'total' in col.lower() or 'cantidad' in col.lower() or \
                   'monto' in col.lower():
                    col_numerica = idx + 1
                    break
        
        if col_numerica is None:
            return
        
        # Crear gráfico de barras
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Resumen"
        
        # Datos
        data = Reference(ws, min_col=col_numerica, min_row=fila_inicio, 
                        max_row=fila_inicio + num_filas)
        cats = Reference(ws, min_col=1, min_row=fila_inicio + 1, 
                        max_row=fila_inicio + min(num_filas, 10))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        
        # Posicionar
        ws.add_chart(chart, f"H{fila_inicio}")
    
    def _agregar_hoja_resumen(self, wb, datos: Dict[str, pd.DataFrame]):
        """Agrega una hoja de resumen con estadísticas."""
        ws = wb.create_sheet(title="Resumen", index=0)
        
        ws.merge_cells('A1:D1')
        ws['A1'] = "RESUMEN EJECUTIVO"
        ws['A1'].font = self.estilos['titulo']
        ws['A1'].fill = self.estilos['fill_header']
        ws['A1'].alignment = self.estilos['center']
        ws.row_dimensions[1].height = 35
        
        fila = 3
        for nombre, df in datos.items():
            ws.cell(row=fila, column=1, value=nombre).font = self.estilos['subtitulo']
            ws.cell(row=fila + 1, column=1, value="Registros:")
            ws.cell(row=fila + 1, column=2, value=len(df))
            
            # Estadísticas de columnas numéricas
            for col in df.select_dtypes(include=['int64', 'float64']).columns:
                ws.cell(row=fila + 2, column=1, value=f"Total {col}:")
                ws.cell(row=fila + 2, column=2, value=df[col].sum())
                ws['B' + str(fila + 2)].number_format = '$#,##0.00'
                fila += 1
            
            fila += 4
    
    # ========================================
    # REPORTES EN PDF
    # ========================================
    
    def crear_pdf_profesional(self, datos: Dict[str, pd.DataFrame],
                              titulo: str = "Reporte",
                              incluir_graficos: bool = True) -> str:
        """
        Crea un PDF profesional con tablas y gráficos.
        
        Args:
            datos: Diccionario {seccion: DataFrame}
            titulo: Título del documento
            incluir_graficos: Si incluir visualizaciones
        
        Returns:
            Ruta del archivo creado
        """
        if not REPORTLAB_DISPONIBLE:
            print("reportlab no instalado. Use: pip install reportlab")
            return self.crear_excel_profesional(datos, titulo, incluir_graficos)
        
        archivo = self._generar_nombre_archivo(titulo.replace(' ', '_'), 'pdf')
        
        doc = SimpleDocTemplate(archivo, pagesize=A4,
                               rightMargin=50, leftMargin=50,
                               topMargin=50, bottomMargin=50)
        
        elementos = []
        estilos = getSampleStyleSheet()
        
        # Estilo de título
        estilo_titulo = ParagraphStyle(
            'TituloPro',
            parent=estilos['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1F4E79'),
            spaceAfter=12,
            alignment=1  # Centro
        )
        
        estilo_subtitulo = ParagraphStyle(
            'SubtituloPro',
            parent=estilos['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2E75B6'),
            spaceBefore=20,
            spaceAfter=10
        )
        
        # Título principal
        elementos.append(Paragraph(titulo, estilo_titulo))
        elementos.append(Paragraph(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            estilos['Normal']
        ))
        elementos.append(Spacer(1, 0.5*inch))
        
        # Procesar cada sección
        for seccion, df in datos.items():
            if df.empty:
                continue
            
            elementos.append(Paragraph(seccion, estilo_subtitulo))
            
            # Limitar filas para PDF
            df_limitado = df.head(50)
            
            # Crear tabla
            tabla_datos = [df_limitado.columns.tolist()]
            for _, fila in df_limitado.iterrows():
                fila_procesada = []
                for val in fila:
                    if isinstance(val, (list, tuple)):
                        val = val[1] if len(val) > 1 else str(val)
                    if isinstance(val, float):
                        val = f"${val:,.2f}"
                    fila_procesada.append(str(val)[:50])  # Truncar texto largo
                tabla_datos.append(fila_procesada)
            
            tabla = Table(tabla_datos)
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F2F2F2')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#CCCCCC')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
                 [colors.white, colors.HexColor('#E8E8E8')]),
            ]))
            
            elementos.append(tabla)
            elementos.append(Spacer(1, 0.3*inch))
            
            # Agregar gráfico si hay datos numéricos
            if incluir_graficos and MATPLOTLIB_DISPONIBLE:
                imagen_grafico = self._crear_grafico_matplotlib(df_limitado, seccion)
                if imagen_grafico:
                    elementos.append(RLImage(imagen_grafico, width=5*inch, height=3*inch))
                    elementos.append(Spacer(1, 0.3*inch))
        
        doc.build(elementos)
        return archivo
    
    def _crear_grafico_matplotlib(self, df: pd.DataFrame, titulo: str) -> Optional[str]:
        """Crea un gráfico con matplotlib y retorna la ruta de la imagen."""
        if not MATPLOTLIB_DISPONIBLE:
            return None
        
        # Buscar columnas numéricas
        cols_numericas = df.select_dtypes(include=['int64', 'float64']).columns
        if len(cols_numericas) == 0:
            return None
        
        col_valor = cols_numericas[0]
        col_etiqueta = df.columns[0]
        
        # Limitar datos para el gráfico
        df_graf = df.head(10)
        
        try:
            fig, ax = plt.subplots(figsize=(8, 5))
            
            # Obtener etiquetas (manejando valores de Odoo)
            etiquetas = df_graf[col_etiqueta].apply(
                lambda x: x[1][:20] if isinstance(x, (list, tuple)) and len(x) > 1 
                else str(x)[:20]
            )
            valores = df_graf[col_valor]
            
            # Crear gráfico de barras
            colores_barras = plt.cm.Blues(range(50, 250, int(200/len(df_graf))))
            ax.barh(etiquetas, valores, color=colores_barras)
            
            ax.set_xlabel(col_valor.replace('_', ' ').title())
            ax.set_title(titulo, fontsize=12, fontweight='bold', color='#1F4E79')
            
            plt.tight_layout()
            
            # Guardar como imagen temporal
            ruta_img = os.path.join(self.directorio, f'temp_chart_{datetime.now().timestamp()}.png')
            plt.savefig(ruta_img, dpi=100, bbox_inches='tight')
            plt.close()
            
            return ruta_img
            
        except Exception as e:
            print(f"Error creando gráfico: {e}")
            return None
    
    # ========================================
    # REPORTES EN HTML
    # ========================================
    
    def crear_html_profesional(self, datos: Dict[str, pd.DataFrame],
                               titulo: str = "Reporte") -> str:
        """Crea un reporte HTML interactivo."""
        archivo = self._generar_nombre_archivo(titulo.replace(' ', '_'), 'html')
        
        html_content = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{titulo}</title>
    <style>
        :root {{
            --primary: #1F4E79;
            --secondary: #2E75B6;
            --accent: #00B050;
            --bg-light: #F5F5F5;
        }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: var(--bg-light);
            color: #333;
            line-height: 1.6;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
        }}
        header {{
            background: linear-gradient(135deg, var(--primary), var(--secondary));
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        h1 {{ font-size: 2.5em; margin-bottom: 10px; }}
        .fecha {{ opacity: 0.8; font-size: 0.9em; }}
        .seccion {{
            background: white;
            border-radius: 10px;
            padding: 20px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }}
        h2 {{
            color: var(--primary);
            border-bottom: 3px solid var(--secondary);
            padding-bottom: 10px;
            margin-bottom: 20px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }}
        th {{
            background: var(--primary);
            color: white;
            padding: 12px 15px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 10px 15px;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{ background: #f8f9fa; }}
        tr:hover {{ background: #e9ecef; }}
        .estadisticas {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 20px;
        }}
        .stat-card {{
            background: linear-gradient(135deg, var(--primary), var(--secondary));
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }}
        .stat-value {{
            font-size: 2em;
            font-weight: bold;
        }}
        .stat-label {{
            opacity: 0.9;
            font-size: 0.9em;
        }}
        .numero {{ text-align: right; font-family: monospace; }}
        .moneda::before {{ content: '$'; }}
        footer {{
            text-align: center;
            padding: 20px;
            color: #666;
            font-size: 0.9em;
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>{titulo}</h1>
            <p class="fecha">Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
        </header>
        
        <div class="estadisticas">
"""
        
        # Agregar tarjetas de estadísticas
        for nombre, df in datos.items():
            total = 0
            if not df.empty:
                for col in df.select_dtypes(include=['int64', 'float64']).columns:
                    if 'total' in col.lower() or 'monto' in col.lower():
                        total = df[col].sum()
                        break
            
            html_content += f"""
            <div class="stat-card">
                <div class="stat-value">{len(df):,}</div>
                <div class="stat-label">{nombre}</div>
                {"<div class='stat-label'>$" + f"{total:,.2f}</div>" if total else ""}
            </div>
"""
        
        html_content += """
        </div>
"""
        
        # Agregar tablas de datos
        for nombre, df in datos.items():
            if df.empty:
                continue
            
            html_content += f"""
        <div class="seccion">
            <h2>{nombre}</h2>
            <table>
                <thead>
                    <tr>
"""
            for col in df.columns:
                html_content += f"                        <th>{col.replace('_', ' ').title()}</th>\n"
            
            html_content += """                    </tr>
                </thead>
                <tbody>
"""
            
            for _, fila in df.head(100).iterrows():
                html_content += "                    <tr>\n"
                for val in fila:
                    if isinstance(val, (list, tuple)):
                        val = val[1] if len(val) > 1 else str(val)
                    if isinstance(val, float):
                        html_content += f'                        <td class="numero moneda">{val:,.2f}</td>\n'
                    else:
                        html_content += f"                        <td>{val}</td>\n"
                html_content += "                    </tr>\n"
            
            html_content += """                </tbody>
            </table>
        </div>
"""
        
        html_content += """
        <footer>
            <p>Generado por ANDROMEDA- Sistema de Análisis Inteligente</p>
        </footer>
    </div>
</body>
</html>
"""
        
        with open(archivo, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return archivo
    
    # ========================================
    # MÉTODO PRINCIPAL
    # ========================================
    
    def generar_reporte(self, datos: Dict[str, pd.DataFrame],
                        titulo: str = "Reporte",
                        formato: str = "excel",
                        incluir_graficos: bool = True) -> str:
        """
        Genera un reporte en el formato especificado.
        
        Args:
            datos: Diccionario {seccion: DataFrame}
            titulo: Título del reporte
            formato: 'excel', 'pdf', 'html' o 'todos'
            incluir_graficos: Si incluir visualizaciones
        
        Returns:
            Ruta del archivo generado (o lista si formato='todos')
        """
        formato = formato.lower()
        
        if formato == 'excel' or formato == 'xlsx':
            return self.crear_excel_profesional(datos, titulo, incluir_graficos)
        
        elif formato == 'pdf':
            return self.crear_pdf_profesional(datos, titulo, incluir_graficos)
        
        elif formato == 'html':
            return self.crear_html_profesional(datos, titulo)
        
        elif formato == 'todos' or formato == 'all':
            archivos = []
            archivos.append(self.crear_excel_profesional(datos, titulo, incluir_graficos))
            archivos.append(self.crear_html_profesional(datos, titulo))
            if REPORTLAB_DISPONIBLE:
                archivos.append(self.crear_pdf_profesional(datos, titulo, incluir_graficos))
            return archivos
        
        else:
            print(f"Formato '{formato}' no reconocido. Usando Excel.")
            return self.crear_excel_profesional(datos, titulo, incluir_graficos)


# ============================================================
# PRUEBAS
# ============================================================

if __name__ == "__main__":
    print("Probando Generador de Reportes...")
    print("=" * 60)
    
    # Datos de prueba
    datos_prueba = {
        'Ventas': pd.DataFrame({
            'Orden': ['SO001', 'SO002', 'SO003'],
            'Cliente': ['Juan', 'María', 'Pedro'],
            'Total': [1500.50, 2300.00, 890.75],
            'Estado': ['Confirmado', 'Pagado', 'Enviado']
        }),
        'Productos': pd.DataFrame({
            'Código': ['PROD001', 'PROD002'],
            'Nombre': ['Casco Premium', 'Guantes Pro'],
            'Stock': [50, 120],
            'Precio': [1299.00, 599.00]
        })
    }
    
    generador = GeneradorReportes()
    
    # Generar Excel
    archivo_excel = generador.crear_excel_profesional(datos_prueba, "Reporte_Prueba")
    print(f"Excel generado: {archivo_excel}")
    
    # Generar HTML
    archivo_html = generador.crear_html_profesional(datos_prueba, "Reporte_Prueba")
    print(f"HTML generado: {archivo_html}")
